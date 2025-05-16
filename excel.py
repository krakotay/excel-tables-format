from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def create_title_sheet(
    output_path: str,
    form_code: str,
    form_number: str,
    report_title: str,
    as_of_date: str,
    formation_date: str,
    number_number: str,
    full_name: str,
    inn: str,
    dob: str,
    pob: str,
    accounts: list[dict],
):
    """
    Создаёт Excel-файл с титульным листом.

    Параметры:
    - output_path: куда сохранить файл, например "test.xlsx"
    - form_code: "Форма по КНД 1120499"
    - form_number: "Форма 67ф"
    - report_title: заголовок вида
        "Сведения о банковских счетах (вкладах, электронных средствах платежа (ЭСП))\nфизического лица, не являющегося ИП"
    - as_of_date: строка даты по состоянию на, например "16.01.2025"
    - formation_date: строка даты формирования, например "16.01.2025"
    - number_number: какой-то номер, например '123'
    - full_name: ФИО
    - inn: ИНН
    - dob: дата рождения, например "27.01.1975"
    - pob: место рождения
    - accounts: список словарей с ключами:
        {
          "bank_name": str,
          "bank_details": str,  # ПерНом/НомФ, ИНН/КПП, БИК
          "address": str,
          "account_number": str,
          "open_date": str,
          "status": str,
          "account_type": str,
        }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Титульник"

    # Стили
    # center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    # Стиль границ
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Шапка формы
    ws.merge_cells("A1:E1")
    ws["A1"] = form_code
    ws["A1"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = form_number
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")

    # Заголовок отчёта
    ws.merge_cells("A3:E4")
    ws["A3"] = report_title
    ws["A3"].font = bold
    ws["A3"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[3].height = 15 * 2


    ws.merge_cells("A6:E6")
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A6"] = CellRichText(
        "По состоянию на ", TextBlock(InlineFont(u="single"), as_of_date)
    )

    # Дата формирования и номер

    # Create rich text with different formatting

    # Merge cells and set alignment
    ws.merge_cells("A8:E8")
    cell = ws["A8"]
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create rich text with different formatting for the number
    rich_text = CellRichText(
        f"Дата формирования {formation_date} ",
        " " * 100,
        "№ ",
        TextBlock(InlineFont(u="single"), str(number_number)),
    )
    cell.value = rich_text

    # ФИО — именно по центру колонки A:E
    ws.merge_cells("A10:E10")
    ws["A10"] = full_name
    ws["A10"].font = bold
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")

    # ИНН — тоже по центру
    ws.merge_cells("A11:E11")
    ws["A11"] = f"ИНН: {inn}"
    ws["A11"].alignment = Alignment(horizontal="center", vertical="center")

    # ДР и МПР — тоже по центру
    ws.merge_cells("A12:E12")
    ws["A12"] = f"Дата рождения: {dob}, Место рождения: {pob}" if dob and pob else ""
    ws["A12"].alignment = Alignment(horizontal="center", vertical="center")

    # Заголовки таблицы
    headers = [
        "Номер счета/номер ЭСП",
        "Дата открытия/предоставления права",
        "Дата закрытия/прекращения права",
        "Состояние",
        "Вид счета",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=col, value=h)
        cell.font = bold
        # Включаем перенос текста и выравнивание по центру
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    # Автоподбор высоты для строки с заголовками
    ws.row_dimensions[14].height = 15  # Устанавливаем начальную высоту
    ws.row_dimensions[14].auto_size = True
    # Увеличиваем высоту на 50% для лучшей читаемости (минимум 20 пикселей)
    current_height = ws.row_dimensions[14].height or 15
    ws.row_dimensions[14].height = max(current_height * 2, 20)

    # Заполняем аккаунты
    row = 14
    for acct in accounts:
        # перед каждой группой выводим блок с названием банка и деталями (без рамки)
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=acct["bank_name"]).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=acct["bank_details"]).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"Адрес: {acct['address']}").alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Строка счёта с рамкой
        row += 1
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

        ws.cell(row=row, column=1, value=acct["account_number"])
        ws.cell(row=row, column=2, value=acct["open_date"])
        ws.cell(row=row, column=3, value=acct.get("close_date", ""))  # можно пустую
        ws.cell(row=row, column=4, value=acct["status"])
        ws.cell(row=row, column=5, value=acct["account_type"])

    # Автоподбор ширины столбцов A-E
    for col in range(1, 6):
        letter = get_column_letter(col)
        max_length = 0
        # Проверяем длину текста в каждой ячейке столбца
        for cell in ws[letter]:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError, ValueError):
                # Skip cells with values that can't be converted to string
                continue
        # Устанавливаем ширину с увеличенным запасом
        adjusted_width = (max_length + 8) * 1.2
        ws.column_dimensions[letter].width = min(
            adjusted_width, 40
        )  # Максимальная ширина 40 символов

    # Применяем границы к заголовкам таблицы (A13:E13)
    for col in range(1, 6):
        ws.cell(row=14, column=col).border = thin_border

    wb.save(output_path)


if __name__ == "__main__":
    accounts = [
        {
            "bank_name": 'Публичное акционерное общество "Сбербанк России", Иркутское отделение № 8586',
            "bank_details": "РегНом/НомФ: 1481/1945 ИНН/КПП: 7707083893/381143001 БИК(СВИФТ): 042520607",
            "address": "664011, г. Иркутск, ул. Пискунова, 122",
            "account_number": "40820810518351951840",
            "open_date": "14.08.2023",
            "status": "открыт",
            "account_type": "Текущий счет",
        },
        # ... другие записи ...
    ]

    create_title_sheet(
        output_path="test.xlsx",
        form_code="Форма по КНД 1120499",
        form_number="Форма 67ф",
        report_title=(
            "Сведения о банковских счетах (вкладах, электронных средствах "
            "платежа (ЭСП))\nфизического лица, не являющегося индивидуальным предпринимателем"
        ),
        as_of_date="16.01.2025",
        formation_date="16.01.2025",
        number_number=123456789101112,
        full_name="ЭРМАТОВ МУСТАФАКУЛ САЙДУЛЛАЕВИЧ",
        inn="384913640328",
        dob="27.01.1975",
        pob="СЫРДАРЬИНСКАЯ ОБЛ",
        accounts=accounts,
    )
